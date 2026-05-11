from __future__ import annotations

import io
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill


HEADERS = [
	"Drawing Number",
	"Title",
	"Company",
	"Revision",
	"Date",
	"Description",
	"Material",
	"Material Source",
	"Dimensions Summary",
	"Min Confidence",
	"Extraction Status",
	"Processed At",
]

FILL_NULL = PatternFill(start_color="FFFF6B6B", end_color="FFFF6B6B", fill_type="solid")
FILL_PARTIAL = PatternFill(start_color="FFFFC857", end_color="FFFFC857", fill_type="solid")


def _first_view_dimensions_summary(drawings_dimensions: Dict[str, Any]) -> str:
	"""Build a short summary string of labeled dimensions from the first view found."""
	views = drawings_dimensions.get("views") or {}
	first_view = next(iter(views.values()), {})
	parts: List[str] = []
	for name, val in sorted((first_view or {}).items()):
		if not isinstance(val, dict):
			continue
		v = val.get("value")
		if v is None:
			continue
		unit = val.get("unit") or ""
		parts.append(f"{name}: {v}{(' ' + unit) if unit else ''}")
	return ", ".join(parts)


def _collect_inferred(dimensions: Dict[str, Any]) -> List[str]:
	"""VLM only extracts labeled dimensions, never infers. Returns empty list."""
	return []


def generate_excel_from_firestore(drawings: List[Dict[str, Any]]) -> bytes:
	"""
	Builds an Excel file as bytes from Firestore drawing documents.
	"""
	wb = Workbook()
	ws = wb.active
	ws.title = "Drawings"
	ws.append(HEADERS)

	for drawing in drawings:
		title = drawing.get("title_block") or {}
		notes = drawing.get("notes") or {}
		dimensions = drawing.get("dimensions") or {}
		summary = drawing.get("extraction_summary") or {}

		material = notes.get("material") or title.get("material")
		material_source = drawing.get("material_source") or "null"

		dims_summary = _first_view_dimensions_summary(dimensions)
		min_confidence = summary.get("min_confidence_score")

		row = [
			title.get("drawing_number"),
			title.get("title"),
			title.get("company"),
			title.get("revision"),
			title.get("date"),
			title.get("title"),
			material,
			material_source,
			dims_summary,
			min_confidence,
			drawing.get("job_status"),
			drawing.get("processed_at"),
		]
		ws.append(row)

		row_idx = ws.max_row
		if drawing.get("job_status") == "partial":
			for col in range(1, len(HEADERS) + 1):
				ws.cell(row=row_idx, column=col).fill = FILL_PARTIAL

		# Highlight entire row for partial extractions; individual dimension highlighting removed
		# (we no longer have fixed per-dimension columns)

	for column_cells in ws.columns:
		max_length = 0
		column = column_cells[0].column_letter
		for cell in column_cells:
			if cell.value is None:
				continue
			max_length = max(max_length, len(str(cell.value)))
		ws.column_dimensions[column].width = min(max(max_length + 2, 12), 50)

	output = io.BytesIO()
	wb.save(output)
	return output.getvalue()
